import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
import excel
from datetime import date

from os.path import exists

from tkinter import *
from PIL import ImageTk, Image

book = openpyxl.load_workbook("BerAmer.xlsx")
root = Tk()
root.title("American Nursery")
root.geometry("300x300")

# Get the first sheet and give it a name
sheet_1 = book.active
sheet_1.title = "Sheet1"

# Production
total_qty_produced = int(sheet_1["C4"].value or 0)

# Costs
total_delivery_costs = int(sheet_1["A12"].value or 0)

total_variable_costs = int(sheet_1["A10"].value or 0)  # Costs that are NOT reoccurring ($R)

total_container_cost = int(sheet_1["A6"].value or 0)  # Sale Containers (Stays with product when sold)

total_medium_cost = int(
    sheet_1["A4"].value or 0)  # R$ A generous estimate of the cost of growing medium is used for business purposes
# Can be revised retrospectively.

total_seed_cost = int(sheet_1["A8"].value or 0)  # R$

total_fixed_costs = total_container_cost + total_medium_cost + total_seed_cost + total_delivery_costs
print(f" Total fixed costs: {total_fixed_costs}")
print(f" Total variable costs: {total_variable_costs}")

total_expenses = int(sheet_1["A12"].value or 0) + int(sheet_1["A10"].value or 0) + int(sheet_1["A8"].value or 0) + int(
    sheet_1["A6"].value or 0) + int(sheet_1["A4"].value or 0)
print(f"\n TOTAL COSTS: {total_expenses}")

#cost_per_plant = total_expenses / int(sheet_1["C4"].value or 0)
# print(f" COST PER PLANT: {cost_per_plant}")

# Sales
total_units_sold = int(sheet_1["E4"].value or 0)
PRICE = 2  # R$

total_revenue = total_units_sold * PRICE

total_profit = total_revenue - total_expenses

#profit_per = total_profit / total_qty_produced

# print(f" PROFIT PER PLANT: {profit_per_plant}")

def update_data():
    global production_entry, sales_entry, medium_buy_entry, container_buy_entry, seed_buy_entry, variable_buy_entry, delivery_buy_entry

    sheet_1["C4"].value = int(production_entry.get()) + int(sheet_1["C4"].value or 0)
    sheet_1["E4"].value = int(sales_entry.get()) + int(sheet_1["E4"].value or 0)
    sheet_1["A4"].value = int(medium_buy_entry.get()) + int(sheet_1["A4"].value or 0)
    sheet_1["A6"].value = int(container_buy_entry.get()) + int(sheet_1["A6"].value or 0)
    sheet_1["A8"].value = int(seed_buy_entry.get()) + int(sheet_1["A8"].value or 0)
    sheet_1["A10"].value = int(variable_buy_entry.get()) + int(sheet_1["A10"].value or 0)
    sheet_1["A12"].value = int(delivery_buy_entry.get()) + int(sheet_1["A12"].value or 0)

    production_entry.delete(0, END)
    sales_entry.delete(0, END)
    medium_buy_entry.delete(0, END)
    container_buy_entry.delete(0, END)
    seed_buy_entry.delete(0, END)
    variable_buy_entry.delete(0, END)
    delivery_buy_entry.delete(0, END)

    book.save("BerAmer.xlsx")

def open_report():
    global production_entry, sales_entry, medium_buy_entry, container_buy_entry, seed_buy_entry, variable_buy_entry, delivery_buy_entry

    report_win = Toplevel()
    report_win.title("Report Event")

    production_entry = Entry(report_win)
    production_entry.grid(row=0, column=1)
    production_label = Label(report_win, text="New Production:")
    production_label.grid(row=0, column=0)

    sales_entry = Entry(report_win)
    sales_entry.grid(row=1, column=1)
    sales_label = Label(report_win, text="Sales:")
    sales_label.grid(row=1, column=0)

    medium_buy_entry = Entry(report_win)
    medium_buy_entry.grid(row=2, column=1)
    medium_label = Label(report_win, text="Medium Purchase:")
    medium_label.grid(row=2, column=0)

    container_buy_entry = Entry(report_win)
    container_buy_entry.grid(row=3, column=1)
    container_label = Label(report_win, text="Container Purchase:")
    container_label.grid(row=3, column=0)

    seed_buy_entry = Entry(report_win)
    seed_buy_entry.grid(row=4, column=1)
    seed_label = Label(report_win, text="Seed Purchase:")
    seed_label.grid(row=4, column=0)

    variable_buy_entry = Entry(report_win)
    variable_buy_entry.grid(row=5, column=1)
    variable_label = Label(report_win, text="Variable Purchase:")
    variable_label.grid(row=5, column=0)

    delivery_buy_entry = Entry(report_win)
    delivery_buy_entry.grid(row=6, column=1)
    delivery_label = Label(report_win, text="Delivery Fee:")
    delivery_label.grid(row=6, column=0)

    # Create 'Submit' button
    sub_butt = Button(report_win, text="Submit", command=update_data)
    sub_butt.grid(row=7, column=1)

    back_butt = Button(report_win, text="Back", command=report_win.destroy)
    back_butt.grid(row=7, column=2)


def open_view():
    view_win = Toplevel()
    view_win.title("Expenses, Revenue, Profit")

    expenses_label = Label(view_win, text="EXPENSES:    ")
    expenses_label.grid(row=0, column=0)
    expenses_num = Label(view_win, text=total_expenses)
    expenses_num.grid(row=0, column=1)

    revenue_label = Label(view_win, text="REVENUE:    ")
    revenue_label.grid(row=1, column=0)
    revenue_num = Label(view_win, text=total_revenue)
    revenue_num.grid(row=1, column=1)

    profit_label = Label(view_win, text="PROFIT:    ")
    profit_label.grid(row=2, column=0)
    profit_num = Label(view_win, text=total_profit)
    profit_num.grid(row=2, column=1)

    profit_per_label = Label(view_win, text="PROFIT/UNIT SOLD:    ")
    profit_per_label.grid(row=3, column=0)
#    profit_per_num = Label(view_win, text=profit_per)
 #   profit_per_num.grid(row=3, column=1)

    back_butt = Button(view_win, text="Back", command=view_win.destroy)
    back_butt.grid(row=4, column=1)


home_label = Label(root, text="Home")
home_label.grid(row=0, column=2)

report_butt = Button(root, text="Report Sale, Production, or Expense", command=open_report)
report_butt.grid(row=1, column=2)

view_butt = Button(root, text="View Expenses, Revenue, and Profit", command=open_view)
view_butt.grid(row=2, column=2)

# if exists("BerAmer.xlsx"):
# Instantiate a workbook from an existing file
#   book = openpyxl.load_workbook("BerAmer.xlsx")

# else:
# Instantiate a workbook
# book = openpyxl.Workbook("BerAmer.xlsx")

book = openpyxl.load_workbook("BerAmer.xlsx")
# Get the first sheet and give it a name
sheet_1 = book.active
sheet_1.title = "Sheet1"

sheet_1["A3"].value = "Medium Costs"
sheet_1["A5"].value = "Container Costs"
sheet_1["A7"].value = "Seed Costs"
sheet_1["A9"].value = "Variable Costs"
sheet_1["A11"].value = "Delivery Costs"
sheet_1["C3"].value = "Units Produced"
sheet_1["E3"].value = "Units Sold"

sheet_1["A1"].value = "Expenses"
sheet_1["A2"].value = total_expenses


sheet_1["C1"].value = "Revenue"
sheet_1["C2"].value = total_revenue

sheet_1["E1"].value = "Profits"
sheet_1["E2"].value = total_profit

sheet_1["G1"].value = "Profit/Unit Sold"
#sheet_1["G2"].value = profit_per
book.save("BerAmer.xlsx")
print("Hellooooooo")
root.mainloop()

# Formatting: fill color, alignment, border, and font.
# font_format = Font(color="FF0000", bold=True)
# thin = Side(border_style="thin", color="FF0000")
# sheet_1["A3"].value = "Hello 3"
# sheet_1["A3"].font = font_format
# sheet_1["A3"].border = Border(top=thin, left=thin,
#                     right=thin, bottom=thin)
# sheet_1["A3"].fill = PatternFill(fgColor="FFFF00", fill_type="solid")

# Number formatting (using Excel's formatting strings)
# sheet_1["A4"].value = 3.3333
# sheet_1["A4"].number_format = "0.00"

# Date formatting (using Excel's formatting strings)
# sheet_1["A5"].value = date(2016, 10, 13)
# sheet_1["A5"].number_format = "mm/dd/yy"

# Formula: Youm must use the English name of the formula with commas as delimiters
# sheet_1["A6"].value = "=SUM(A4, 2)"

# Image


# Two-dimensional list (we're using our Excel module)
# data = [[None, "North", "South"],
#      ["Last year", 2, 5],
#     ["This year", 3, 6]]
# excel.write(sheet, data, "A10")

# Chart
# chart = BarChart()
# chart.type = "col"
# chart.title = "Sales Per Region"
# chart.x_axis.title = "Regions"
# chart.y_axis.title = "Sales"
# chart_data = Reference(sheet, min_row=11, min_col=1,
#                    max_row=12, max_col=3)
# chart_categories = Reference(sheet, min_row=10, min_col=2,
#                     max_row=10, max_col=3)
# "from_rows" interprets the data in the same way as if you would add a chart manually in Excel
# chart.add_data(chart_data, titles_from_data=True, from_rows=True)
# chart.set_categories(chart_categories)
# sheet.add_chart(chart, "A15")

